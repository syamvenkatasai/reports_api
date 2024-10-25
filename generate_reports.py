from ace_logger import Logging
from db_utils import DB
from jinja2 import Environment
from openpyxl.styles import Border, Side, Alignment
from premailer import Premailer
from openpyxl import load_workbook
from pathlib import Path
from pandas import json_normalize

from datetime import datetime
from copy import copy
from tablepyxl import tablepyxl
from tablepyxl.tablepyxl import get_Tables, write_rows
from jinja2 import Template
import openpyxl
import traceback
import base64
import pytz
import subprocess
import pandas as pd
import time
import requests
import json
import os
import urllib.parse
import re

logging = Logging(name='reports_consumer')
db_config = {
                    'host': os.environ['HOST_IP'],
                    'password': os.environ['LOCAL_DB_PASSWORD'],
                    'user': os.environ['LOCAL_DB_USER'],
                    'port': os.environ['LOCAL_DB_PORT'],
                }

def reports_consumer(data):
    try:
        print(f'Message: {data}')
        reference_id = data.get('REFERENCE_ID', None)
        tenant_id = data.get('tenant_id', None)

        if reference_id is None:
           logging.debug('Recieved wrong data. Commiting.')
           return {"flag":False,"message":"something went wrong!! reference_id is none in reports_conumer"}
        try:
            logging.info("Entered into reports_consumer")
            generate_report_reports_consumer(data)
        except Exception:
            logging.exception("something went wrong in reports_consumer!!!")
            db_config['tenant_id'] = tenant_id
            reports_db = DB('reports', **db_config)
            query = 'UPDATE reports_queue SET status=%s WHERE reference_id=%s'
            reports_db.execute_(query, params=['Failed', reference_id])
    except Exception:
        logging.exception('Something went wrong in consumer. Check trace.')

def generate_report_reports_consumer(data):
    logging.info("generate_report_reports_consumer request data: ",data)
    tenant_id = data['tenant_id']
    reference_id = data['REFERENCE_ID']
    query_type=data['QUERY_TYPE']
    report_id=int(data["REPORT_ID"])
    parent_id = int(data['parent_id'])
    parent_ref_id = data['parent_ref_id']
    
    db_config['tenant_id'] = tenant_id
    reports_db = DB('reports', **db_config)

    query = f"SELECT * FROM report_requests WHERE reference_id='{reference_id}'"
    report_type_df = reports_db.execute_(query)

    report_type_query=""
    report_data=[]
    generate_report_flag=False

    ### Checking if Jinja template exists to know whether to generate an df to excel or jinjia to excel
    template_check_query = f"SELECT report_template_jinja FROM report_template WHERE report_id={report_id}"
    jinja_data_list = reports_db.execute_(template_check_query).to_dict('records')

    logging.info("############## JINJA DATA LIST LENGTH {len(jinja_data_list)}",len(jinja_data_list))

    report_query = report_type_df['actual_report_query'].to_list()
    if len(report_query)>0:
        report_type_query=report_type_df['actual_report_query'][0]
        logging.info(f"############ Query Type {query_type} and ACTUAL Query to RUN: {report_type_query}")
    
        if query_type=="query":
            try:

                ### check if there escape single quotes replace them with single quote
                if "''" in report_type_query:
                    report_type_query=report_type_query.replace("''","'")
                    
                query_db1 = DB("group_access", **db_config)
                report_data = query_db1.execute_(report_type_query).to_dict('records')
                reports_json = {}
                reports_json["report_data"]=report_data
                reports_json['ui_data']=data
                reports_json['tenant_id'] = tenant_id

                ################# TO DO: PRODUCE TO GENERATE Report ##############
                generate_report({'tenant_id': tenant_id, **reports_json})
                
                generate_report_flag=True
                return 0
            except Exception as e:
                logging.exception("########### Error in quering generate report")
                
        elif query_type=="route":
            query_json=json.loads(report_type_df['query_params'].to_list()[0])
            query_json['tenant_id']=tenant_id
            query_json['ui_data']=data

            request_timeout= 60
            query_json['request_timeout']=request_timeout
            try:
                host_port = "443"
                host_ip = 'reportsapi'
                base_route = f"https://{host_ip}:{host_port}/"
                final_url = base_route+report_type_query

                logging.info("########### Executing route: ",final_url)
                logging.info("########### Executing query_json: ",query_json)
                query_json['ui_data']['REQUESTED_DATETIME'] = query_json['ui_data']['REQUESTED_DATETIME'].strftime('%Y-%m-%d %H:%M:%S')
                query_json['ui_data']['ETA'] = query_json['ui_data']['ETA'].strftime('%Y-%m-%d %H:%M:%S')
                logging.info(f"################# after dateformat: {query_json}")
                response = requests.post(final_url,json=query_json, timeout=request_timeout ,stream=True,verify=False)
                logging.info("################  response: ",response)
                logging.info("################  type response: ",type(response))
                logging.info("################  response_text: ",response.text)
                response_text = response.text.replace('NaN', 'null')
                try:

                    response_data=response_text.json()
                except Exception as e:
                    response_data = json.loads(response_text)
                
                if 'audit' in report_type_query:
                    new_row_data = []
                    try:
                        for row_item in response_data['data'][0]['row_data']:


                            new_row_item = {
                                "S_No":row_item.get('serial_number', ''),
                                "case_id": row_item.get('case_id', ''),
                                "Party_Id" : row_item.get('Party_id', ''),
                                "Party_Name" : row_item.get('Party_name', ''),
                                
                                'Business_Rules_Validation' : str(row_item.get('Business Rules Validation', '')),
                                "CLIMS_Create_API_Request_time" : str(row_item.get('CLIMS Create API Request time', '')),
                                
                                'CLIMS_Create_API_Response_Time' : str(row_item.get('CLIMS Create API Response Time', '')),
                                "Case_creation_time_stamp" : str(row_item.get('Case creation time stamp', '')),
                                "Maker_queue_in_Time_Stamp" : str(row_item.get('Maker queue in Time Stamp', '')),
                                
                                "Completed_queue_in_time_Stamp" : str(row_item.get('Completed queue in time Stamp', '')),
                                "Rejected_queue_in_time_Stamp" : str(row_item.get('Rejected queue in time Stamp', '')),
                                "Total_Handling_time" : str(row_item.get('Total_Handling_time', '')),
                                "Maker_Name" : str(row_item.get('Maker Name', '')),
                                "User_ID" : str(row_item.get('User ID', '')),

                            }
                            new_row_data.append(new_row_item)

                        # Update response_data with the modified row_data
                        response_data['data'][0]['row_data'] = new_row_data
                    except Exception as e:
                        logging.info("#######eeption : ",e)

                logging.info("################  response_data: ",response_data)
                generate_report_flag=response_data['flag']
                logging.debug("generate_report_flag :",generate_report_flag)
                
                if generate_report_flag:
                    producer_data={}
                    producer_data['ui_data']=data
                    producer_data["report_data"]=response_data
                    producer_data['tenant_id'] = tenant_id
                    logging.debug(f"producer_data :{producer_data}")
                    ################# TO DO: PRODUCE TO GENERATE Report ##############
                    generate_report({'tenant_id': tenant_id, **producer_data})
                    
                    return 0
                else:
                    print("################ Route Returned Fail Case updating the status as failed")
                    query = 'UPDATE report_requests SET status=%s WHERE reference_id=%s'
                    if parent_id == -1:
                        reports_db.execute_(query, params=['Failed',reference_id])
                    else:
                        reports_db.execute_(query, params=['Failed',parent_ref_id])
                    return 0

            except Exception as e:
                print("########## something went wrong: ",report_type_query)
                logging.exception(e)
                print("########## something went wrong: ",e)
                print("######## Making the Route Calling Async through route and calling generate report")
                return 0
    else:
        print("################ ERROR No actual query found")
        return 0


def set_for_keys(my_dict, key_arr, val):
    """
    Set val at path in my_dict defined by the string (or serializable object) array key_arr
    """
    current = my_dict
    for i in range(len(key_arr)):
        key = key_arr[i]
        if key not in current:
            if i == len(key_arr) - 1:
                current[key] = val
            else:
                current[key] = {}
        else:
            if type(current[key]) is not dict:
                print("Given dictionary is not compatible with key structure requested")
                raise ValueError("Dictionary key already occupied")

        current = current[key]

    return my_dict

def to_formatted_json(df, sep="."):
    result = []
    for _, row in df.iterrows():
        parsed_row = {}
        for idx, val in row.iteritems():
            keys = idx.split(sep)
            parsed_row = set_for_keys(parsed_row, keys, val)

        result.append(parsed_row)
    return result

def compress(json):
    result = []

    for json_item in json:
        if len(result) == 0:
            result.append(json_item)

        else:
            found = False
            if not found:
                result.append(json_item)
    if len(result) == 1:
        return result[0]
    else:
        return result

def document_to_one_sheet_workbook(doc):
    #Initiating the workbook and the worksheet
    wb = tablepyxl.Workbook()
    sheet = wb.active

    inline_styles_doc = Premailer(doc, remove_classes=False,align_floating_images=False).transform()
    tables = tablepyxl.get_Tables(inline_styles_doc)
    #Appending tables one below the other without leaving extra rows.
    #To append column wise, use write_columns
    
    col=1
    for table in tables:
        if table.head:
            tablepyxl. write_rows(sheet, table.head,1,column =col)
        if table.body:
            tablepyxl. write_rows(sheet, table.body, 1,column =col)
        col+=5
    return wb

def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        row.alignment = alignment
        if fill:
            for c in row:
                c.fill = fill


def generate_report_file(tenant_id, report_id, data, reference_id,file_name_var, sheet_dict, excel_flag):
    try:
        logging.info(f"report_id is {report_id}, In generate_report_file  data recieved is {data}")
        if report_id == 21:
            for row in data[0]['row_data']:
                for key, value in row.items():
                    if isinstance(value, str):
                        if "''" in value:
                            row[key] = value.replace("''", '\"')
                        elif "'" in value:
                            row[key] = value.replace("'", '\"')
        elif report_id in (105,106,107,81):
            for row in data['row_data'][0]:
                for key, value in row.items():
                    if isinstance(value, str):
                        if "''" in value:
                            row[key] = value.replace("''", '\"')
                        elif "'" in value:
                            row[key] = value.replace("'", '\"')
        
        logging.info(f"In generate_report_file after filter data recieved is {data}")
        
        db_config['tenant_id'] = tenant_id
        db = DB('reports', **db_config)
        query = 'SELECT * FROM `report_template` WHERE `report_id`=%s'
        report_data = db.execute_(query, params=[report_id]).to_dict('records')[0]
        logging.info("############### Received Report Data")
        filename = file_name_var
        file_type = report_data["report_out_format"]
        report_template = urllib.parse.unquote(report_data["report_template_jinja"])
        logging.info(f"######## report template is {report_template}")
        logging.info(f"######## report_id is {report_id} and type is {type(report_id)}")
        if report_id in [61 ,41,42,43]:
            template = Template(report_template)
            rendered_html = template.render(row_data=data['row_data'])

            # Store the rendered HTML in a variable
            html_out = rendered_html

        else:
            if report_id != 81 and report_id != 122 and report_id != 105 and report_id != 106 and report_id != 107 and report_id != 103 and report_id != 9999 and report_id != 21 and report_id !=28 and report_id !=29 and report_id !=141:
                df = pd.DataFrame({'row_data': [[]]})
            else:
                df = pd.DataFrame(data)

            logging.info(f"#dataframe for data is {df}")
            # below lines may change according to the data how we get
            try:
                json_df = to_formatted_json(df)
                nested_df = json.dumps((compress(json_df)), ensure_ascii=False)
                data = nested_df.replace("'", "\"")
            except Exception as e:
                #handling data format error (Object of type 'Timestamp' is not JSON serializable)
                logging.info(f"in exception {e}")
                json_df = df.to_json(orient="records", date_format="iso", date_unit="s")
                nested_df = json.dumps(json.loads(json_df), ensure_ascii=False).replace("'", "\"")
                data = nested_df
            logging.debug("#############3data: ", data)

            jinja2_env = Environment()
            print("report_template for jinja:", report_template)
            jinja2_tpl = jinja2_env.from_string(report_template)
            print("report_template for jinja2_tpl: ", jinja2_tpl)
            
            data = json.loads(data)
            logging.debug("###############$$$$data:", data)
            try:
                html_out = jinja2_tpl.render(data)
            except Exception as e:
                #error handling for the data which is coming in row_data format
                row_data = data[0]['row_data']
                html_out = jinja2_tpl.render(row_data=row_data)
            print("html_out***************:", html_out)
            print("debugging file type ", file_type)
            print("############ Received REPORT HTML")
            if (file_type == 'xlsx'):
                file_path = Path(f'./reports/{filename}')
                logging.debug(file_path)
                
                if sheet_dict == -1 and excel_flag == -1:
                    print("#########Document to one sheet workbook")
                    wb = document_to_one_sheet_workbook(html_out)
                    ws = wb['Sheet']
                    thin = Side(border_style="thin", color="000000")
                    border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    al = Alignment(vertical="center",horizontal='center', wrap_text = True)

                    for mer_Cel in ws.merged_cells.ranges:
                        style_range(ws, str(mer_Cel), border=border, alignment=al)

                    wb.save(file_path)
                elif excel_flag == 1:
                    print("####################Tablepyxl Excel flag ")
                    try:
                        wb = document_to_one_sheet_workbook(html_out)
                        ws = wb['Sheet']
                        thin = Side(border_style="thin", color="000000")
                        border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        al = Alignment(vertical="center",horizontal='center', wrap_text = True)
                        for mer_Cel in ws.merged_cells.ranges:
                            style_range(ws, str(mer_Cel), border=border, alignment=al)

                        wb.save(file_path)
                    except Exception as e:
                        logging.exception("something went wrong in tablepyxl excel",e)
                    
                else:
                    wb = tablepyxl.document_to_workbook(html_out)
                    ws = wb['Sheet']
                    thin = Side(border_style="thin", color="000000")
                    border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    al = Alignment(vertical="center",horizontal='center', wrap_text = True)

                    for mer_Cel in ws.merged_cells.ranges:
                        style_range(ws, str(mer_Cel), border=border, alignment=al)

                    wb.save(file_path)

                    workbook = load_workbook(file_path)
                    sheet_names = workbook.sheetnames

                    for sheet in sheet_names:
                        spread_sheet = workbook[sheet]
                        spread_sheet.title = sheet_dict[sheet]
                    workbook.save(file_path)
                    


            elif (file_type=='docx'):
                print("debugging Docx format report")
                filename = filename[:-5]
                saving_temp_html_file = open(Path(f'./reports/{filename}.html'),"w+")
                print(f"debugging html file saved with filename {filename}")
                saving_temp_html_file.write(html_out)
                saving_temp_html_file.close()
                print("debugging html file is created")
                file_path = f'/var/www/reports_api/reports/{filename}.html'
                out_path = '/var/www/reports_api/reports'
                os.system('soffice --invisible --convert-to docx:"MS Word 2007 XML" ' + "'"+str(file_path) + "'"+' --outdir ' + out_path)
                print("running Subprocess")
                subprocess.run("sudo chmod -R +777 ./reports", shell=True)
                query = 'soffice --invisible --convert-to docx:"MS Word 2007 XML" ' + "'"+str(file_path)+"'"+ ' --outdir '+ out_path
                subprocess.run(query, shell=True)
                print(f'The path used for html document => {str(file_path)}')
                print("debugging docx is generated")
            print("updating html report to DB")
        
        try:
            query = "UPDATE REPORT_REQUESTS SET HTML_REPORT= %s WHERE REFERENCE_ID=%s "
            db.execute_(query, params=[html_out,reference_id])
        except Exception as e:
            print("############HTML LENGTH IS HUGE")
            num_tables = [m.start() for m in re.finditer('<table', html_out)]
            how_many_rows = 5
            start_ind = 0
            table_start = 0
            html = ""
            for idx in range(len(num_tables)):
                start_ind = num_tables[idx]
                if idx != len(num_tables) - 1:
                    tr_oc = [m.start() for m in re.finditer('<tr', html_out[table_start:num_tables[idx + 1]])]
                    try:
                        till_last = num_tables[idx] + tr_oc[how_many_rows] - 1
                    except Exception as e:
                        till_last = num_tables[idx] + tr_oc[1] - 1
                    sub_html = html_out[start_ind:till_last]
                else:
                    tr_oc = [m.start() for m in re.finditer('<tr', html_out[num_tables[idx]:])]
                    try:
                        till_last = num_tables[idx] + tr_oc[how_many_rows] - 1
                    except Exception as e:
                        till_last = num_tables[idx] + tr_oc[1] - 1
                    sub_html = html_out[start_ind:till_last]

                html = html + sub_html + '</table>'
                html = re.sub('\n', '', html)
                query = "UPDATE REPORT_REQUESTS SET HTML_REPORT= %s WHERE REFERENCE_ID=%s "
                db.execute_(query, params=[html,reference_id])
    except Exception as e:
        logging.info("something went wrong in generate report file ", e)


def generate_report(ui_data):
    print(f"################--------------ui_data is ------------->{ui_data}")
    data=ui_data['ui_data']
    tenant_id = data['tenant_id']
    file_name = data['REPORT_OUTPUT']
    reference_id = data['REFERENCE_ID']
    query_type=data['QUERY_TYPE']
    report_id=int(data["REPORT_ID"])
    parent_id = int(data['parent_id'])
    parent_req_id = int(data['parent_req_id'])
    parent_ref_id = data['parent_ref_id']
    fund_name = data['fund_name']
    
    db_config['tenant_id'] = tenant_id
    reports_db = DB('reports', **db_config)

    start_time = time.time()

    # Get the report type configuration
    query = f"SELECT * FROM report_requests WHERE reference_id='{reference_id}'"
    print(f"####################### Fetching everything Query: {query}")
    report_type_df = reports_db.execute_(query)

    # Get the report data using query or function
    report_type_query=""
    report_to_excel_flag=False
    report_data=[]
    generate_report_flag=False

    ### Checking if Jinja template exists to know whether to generate an df to excel or jinjia to excel
    template_check_query = f"SELECT report_template_jinja FROM report_template WHERE report_id={report_id}"
    jinja_data_list = reports_db.execute_(template_check_query).to_dict('records')

    print(f"############## JINJA DATA LIST LENTH {len(jinja_data_list)}")

    if len(jinja_data_list)>0:
        jinja_data = jinja_data_list[0]['report_template_jinja']
        if jinja_data=="" or jinja_data==None:
            report_to_excel_flag=True
    else:
        report_to_excel_flag=True
    
    report_query = report_type_df['actual_report_query'].to_list()
    if len(report_query)>0:
        report_type_query=report_type_df['actual_report_query'][0]
        print(f"############ Query Type {query_type} and ACTUAL Query to RUN: {report_type_query}")
        
        if query_type=="query":
            try:                
                report_data=ui_data["report_data"]
                report_data1 = {}                              
                report_data1['row_data'] = [report_data]
                report_data = report_data1
                print(f"$$$$$$$$$5%%%%%%%:{report_data}")

                sheet_dict = -1
                excel_flag = -1
                
                generate_report_flag=True
            except Exception as e:
                print("########### Error in quering generate report")
                print(e)
        elif query_type=="route":
            query_json=json.loads(report_type_df['query_params'].to_list()[0])
            query_json['tenant_id']=tenant_id
            try:
                response_data = ui_data.get("report_data",'')
                generate_report_flag=response_data['flag']
                report_file_data_path = response_data.get('file_path','')
                logging.debug(f"file path received is {report_file_data_path}")
                try:
                    if report_file_data_path:
                        logging.debug(f"reading file to get data")
                        with open(report_file_data_path,'r') as file: 
                            file_json_data = file.read()
                            file_json_data = file_json_data.replace("'","")
                        report_data = json.loads(file_json_data)
                        report_data = [report_data]
                        logging.debug(f"length and type of report data is '{len(report_data)} , '{type(report_data)}' ")
                        logging.debug(f"Report Data is {report_data}")
                    else:
                        logging.debug(f"length and type of  normal report data is '{len(report_data)} , '{type(report_data)}' ")
                        report_data=response_data['data']
                        print(f"=========================<<<<<<<<<<{report_data}")
                except Exception as e:
                    logging.exception(e)
                    traceback.print_exc()

                response_tags=response_data.get('tags',[])
                print(f"response_tags:{response_tags}")
                time_taken = float(response_data.get('time_taken',1))
                sheet_dict=response_data.get('sheet_dict',-1)
                excel_flag =response_data.get('excel_flag',-1)
                if len(response_tags)>0:
                    response_tags=json.dumps(response_tags)
                    update_query = f"UPDATE report_requests SET tags='{response_tags}' WHERE reference_id='{reference_id}'"
                    reports_db.execute_(update_query)
            except Exception as e:
                logging.exception("################ Error in route")
    else:
        logging.info("################ ERROR No actual query found")
    
    if generate_report_flag:
        # if report_to_excel_flag:
        if report_to_excel_flag and report_data:
            try:
                report_data=pd.DataFrame.from_dict(json_normalize(report_data), orient='columns')
            except Exception as e:
                report_data=pd.DataFrame.from_dict(report_data, orient='columns')
                
            writer = pd.ExcelWriter(f'./reports/{file_name}', engine='xlsxwriter')
            logging.info(f'###file_name {file_name}')
            try:
                query = f'SELECT * FROM REPORT_TEMPLATE WHERE REPORT_ID= {report_id}'
                report_data = reports_db.execute_(query).to_dict('records')[0]
                report_template = urllib.parse.unquote(report_data["report_template_jinja"])
                print(f"###report_data {report_data}")
                df = pd.DataFrame(report_data)
                json_df = to_formatted_json(df)
                print(f"###json_df {json_df}")
                nested_df = json.dumps((compress(json_df)), ensure_ascii=False)
                data = nested_df.replace("'", "\"")
                print(f"#############3data:{data}")
                jinja2_env = Environment()
                print(f"report_template for jinja: {report_template}")
                jinja2_tpl = jinja2_env.from_string(report_template)
                print(f"report_template for jinja2_tpl: {jinja2_tpl}")
                data = json.loads(data)
                print(f"###############$$$$data:{data}")

                html_out = jinja2_tpl.render(data)
                print(f"html_out***************:{html_out}")
            except Exception as e:
                logging.info(f"exception while generating preview##  {e}")
            
            try:
                query = f"UPDATE report_requests SET html_report='{html_out}' WHERE reference_id='{reference_id}'"
                reports_db.execute_(query)
            except Exception as e:
                print("############HTML LENGTH IS HUGE")
                num_tables = [m.start() for m in re.finditer('<table', html_out)]
                how_many_rows = 5
                start_ind = 0
                table_start = 0
                html = ""
                for idx in range(len(num_tables)):
                    start_ind = num_tables[idx]
                    if idx != len(num_tables) - 1:
                        tr_oc = [m.start() for m in re.finditer('<tr', html_out[table_start:num_tables[idx + 1]])]
                        try:
                            till_last = num_tables[idx] + tr_oc[how_many_rows] - 1
                        except Exception as e:
                            till_last = num_tables[idx] + tr_oc[1] - 1
                        sub_html = html_out[start_ind:till_last]
                    else:
                        tr_oc = [m.start() for m in re.finditer('<tr', html_out[num_tables[idx]:])]
                        try:
                            till_last = num_tables[idx] + tr_oc[how_many_rows] - 1
                        except Exception as e:
                            till_last = num_tables[idx] + tr_oc[1] - 1
                        sub_html = html_out[start_ind:till_last]

                    html = html + sub_html + '</table>'
                    html = re.sub('\n', '', html)

                query = f"UPDATE report_requests SET html_report='{html}' WHERE reference_id='{reference_id}'"
                reports_db.execute_(query)


            if 'audit' in file_name.lower():
                report_data = pd.DataFrame.from_records(report_data['row_data'][0])
            report_data.to_excel(writer, index=False)
            writer.save()
        elif report_data:
            ### Generating Report through jinja template
            try:
                report_data = report_data[0]['data']
                l=[]
                l.append(report_data['row_data'])
                report_data['row_data']=l
            except Exception as e:
                report_data = report_data
            generate_report_file(tenant_id, report_id, report_data, reference_id,file_name, sheet_dict, excel_flag)
            logging.info('Reported generated')
        else:
            ### in the route itself , the excel file is getting generated so , no need to perform action 
            pass
        ist = pytz.timezone("Asia/Calcutta")
        timestamp = datetime.now(ist)
        
        timestamp = str(timestamp)[:-13]

        took_time = (time.time() - start_time)
        if parent_id != -1:
            print("#################################Child status")
            formatted_requested_time = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')

            child_query = f"UPDATE report_requests SET status='Download',generated_datetime='{formatted_requested_time}',process_time={took_time} WHERE reference_id='{reference_id}'"
            reports_db.execute_(child_query)
            print("#######################child status updated")
            child_status_query = f"SELECT status, requested_datetime FROM report_requests where parent_id = '{parent_req_id}'"
            childrens_status = reports_db.execute_(child_status_query)
            status_list = list(childrens_status['status'])
            print(f"###############Status list {status_list} and {len(status_list)} and {status_list.count('Download')}")
            status_flag = (status_list.count('Download') == len(status_list))
            print("##########################Childrens count")
            print(f"#######################Status Flag {status_flag}")
            if status_flag == True:
                print("#############status flag is True")
                parent_ref_query = f"SELECT request_id FROM report_requests where reference_id = '{parent_ref_id}'"
                parent_req_id = reports_db.execute_(parent_ref_query).values.tolist()[0][0]
                print("#########################report id with parent req reference id")
                child_ref_ids = f"SELECT rr.report_id, rr.reference_id, rt.report_sheetname FROM report_requests as rr left join report_template as rt on rr.report_id = rt.report_id where parent_id = '{parent_req_id}'"
                child_ids = reports_db.execute_(child_ref_ids).to_dict(orient="records")
                print("###############sheet names")
                filename_filter = os.listdir('./reports/')
                print("Debugging file stitching beginning")
                files_stitch = {}
                for idx in range(len(filename_filter)):
                    for j in range(len(child_ids)):
                        try:
                            path_ref_id = filename_filter[idx].split('#')[1].split('.')[0]
                            if child_ids[j]['reference_id'] == path_ref_id:
                                print(child_ids[j]['report_sheetname'])
                                files_stitch[child_ids[j]['report_sheetname']] = '/app/reports/'+filename_filter[idx]

                        except Exception as e:
                            pass

                print("Debugging fetched all DB data")
                parent_name_query = f"SELECT report_name FROM report_requests where reference_id = '{parent_ref_id}'"
                parent_name = reports_db.execute_(parent_name_query).values.tolist()[0][0]
                file_time = files_stitch[list(files_stitch.keys())[0]]
                file_time = file_time.split('-', 1)[1].split('#')[0]
                parent_file_name = f"{parent_name}-{fund_name}-{file_time}.xlsx"
                stitch_file_name = Path(f"/app/reports/{parent_file_name}")
                print(f"Parent file is created at {stitch_file_name}")
                wb = openpyxl.Workbook()

                new_sheet = wb.active

                check = 0
                for fsheet in files_stitch.keys():

                    if check == 0:
                        new_sheet.title = fsheet

                    workbook = load_workbook(files_stitch[fsheet])
                    default_sheet = workbook['Sheet']

                    if check > 0:
                        new_sheet = wb.create_sheet(index=check + 1, title=fsheet)

                    for idx in range(len(list(default_sheet.merged_cells))):
                        new_sheet.merge_cells(str(list(default_sheet.merged_cells)[idx]))
                    for row in default_sheet.rows:
                        for cell in row:
                            none_check = getattr(cell, "value", None)
                            if none_check == None:
                                new_cell = new_sheet.cell(row=cell.row, column=cell.column)
                                if cell.has_style:
                                    new_cell.font = copy(cell.font)
                                    new_cell.border = copy(cell.border)
                                    new_cell.fill = copy(cell.fill)
                                    new_cell.number_format = copy(cell.number_format)
                                    new_cell.protection = copy(cell.protection)
                                    new_cell.alignment = copy(cell.alignment)
                                continue
                            new_cell = new_sheet.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                            if cell.has_style:
                                new_cell.font = copy(cell.font)
                                new_cell.border = copy(cell.border)
                                new_cell.fill = copy(cell.fill)
                                new_cell.number_format = copy(cell.number_format)
                                new_cell.protection = copy(cell.protection)
                                new_cell.alignment = copy(cell.alignment)

                    check = check + 1

                wb.save(stitch_file_name)
                print("Parent file is saved")
                took_time = (datetime.now() - min(list(childrens_status['requested_datetime']))).seconds
                print("Debugging all status location")
                query = f"UPDATE report_requests SET status='Download',generated_datetime='{timestamp}',process_time={took_time},report_output='{parent_file_name}' WHERE reference_id='{parent_ref_id}'"
                reports_db.execute_(query)
                
        if parent_id == -1:
            ist = pytz.timezone("Asia/Calcutta")
            timestamp = datetime.now(ist)
            formatted_datetime = timestamp.strftime('%Y-%m-%d %H:%M:%S')
            logging.info(f"###formatted_requested_time {formatted_datetime}")
            query = f"UPDATE report_requests SET status='Download',generated_datetime=TO_DATE('{formatted_datetime}', 'YYYY-MM-DD HH24:MI:SS'),process_time={took_time} WHERE reference_id='{reference_id}'"
            reports_db.execute_(query)

    else:
        query = 'UPDATE report_requests SET status=%s WHERE reference_id=%s'
        if parent_id == -1:
            reports_db.execute_(query, params=['Failed',reference_id])
        else:
            reports_db.execute_(query, params=['Failed',parent_ref_id])

